import pandas as pd
from pathlib import Path
from tkinter import Tk, filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import os

def pick_file(prompt):
    print(prompt)
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    root.destroy()
    return file_path

def main():
    print("\n=== Side-by-Side Excel Highlighter ===\n")

    # Pick CSVs
    file1 = pick_file("Select first CSV file")
    file2 = pick_file("Select second CSV file")

    # Load data
    df1 = pd.read_csv(file1, dtype=str).fillna("")
    df2 = pd.read_csv(file2, dtype=str).fillna("")

    print("\nColumns in first file:", list(df1.columns))
    col1 = input("Enter the column name to use from File 1: ").strip()
    col2 = input("Enter the column name to use from File 2: ").strip()

    vals1 = df1[col1].dropna().unique().tolist()
    vals2 = df2[col2].dropna().unique().tolist()

    # Ask output name
    output_name = input("\nEnter output file name (without extension): ").strip()
    if not output_name:
        output_name = "side_by_side_comparison"

    # File names for headers
    file1_name = os.path.splitext(os.path.basename(file1))[0]
    file2_name = os.path.splitext(os.path.basename(file2))[0]

    # Create workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Side_by_Side"

    # Headers
    ws1["A1"] = f"{file1_name} ({col1})"
    ws1["B1"] = f"{file2_name} ({col2})"

    # Fill values
    for i, val in enumerate(vals1, start=2):
        ws1[f"A{i}"] = val
    for i, val in enumerate(vals2, start=2):
        ws1[f"B{i}"] = val

    # Conditional formatting
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws1.conditional_formatting.add("A2:A5000",
        FormulaRule(formula=['COUNTIF(B:B,A2)>0'], fill=green_fill))
    ws1.conditional_formatting.add("B2:B5000",
        FormulaRule(formula=['COUNTIF(A:A,B2)>0'], fill=green_fill))
    ws1.conditional_formatting.add("A2:A5000",
        FormulaRule(formula=['AND(A2<>"",COUNTIF(B:B,A2)=0)'], fill=red_fill))
    ws1.conditional_formatting.add("B2:B5000",
        FormulaRule(formula=['AND(B2<>"",COUNTIF(A:A,B2)=0)'], fill=red_fill))

    # Precompute unique values
    only_in_file1 = sorted(set(vals1) - set(vals2))
    only_in_file2 = sorted(set(vals2) - set(vals1))
    all_unique = only_in_file1 + only_in_file2

    # Sheet 2
    ws2 = wb.create_sheet("Unique_Values")
    ws2["A1"] = "Values only in one file"
    for i, val in enumerate(all_unique, start=2):
        ws2[f"A{i}"] = val

    # Save
    output_path = Path.home() / "Downloads" / f"{output_name}.xlsx"
    wb.save(output_path)

    print(f"\n✅ Excel created with conditional formatting & unique list. danieliscool:\n{output_path}")

if __name__ == "__main__":
    main()
